from django.shortcuts import render
from openpyxl import load_workbook
from .models import Product
from .serializer import ProductSerializer
from rest_framework import generics
from django.core.exceptions import ValidationError
from .validators import validate_file_extension
def import_from_excel(request):
    context = {}
    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            validate_file_extension(excel_file)
            wb = load_workbook(excel_file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                id,name,description,location,price,color = row
                Product.objects.create(id=id,name=name,description=description,
                                    location=location,price=price,color=color)

            context['success'] = 'Successful import!'
        except ValidationError as ve:
            context['error'] = f'Error in import: {str(ve)}'
        except Exception as e:
            context['error'] = f'Error in import: {str(e)}'

    return render(request, 'form.html', context)

class ProductListCreateView(generics.ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer